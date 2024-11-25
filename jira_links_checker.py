import os

# Установка необходимых библиотек
os.system("pip install jira")
os.system("pip install openpyxl")
os.system("pip install selenium")
os.system("pip install webdriver-manager")

import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager

input("Для запуска программы нажмите Enter")

# Путь к Вашему Excel-файлу
exl_path = r"jira.xlsx"
exl = load_workbook(exl_path)

# Работа с листом data
sheet_data = exl["data"]

login_exl = sheet_data.cell(row=2, column=1).value  # Логин
password_exl = sheet_data.cell(row=2, column=2).value  # Пароль

# Работа с листом jira
sheet_jr = exl["jr"]

# Инициализация WebDriver
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# Вход на сайт
driver.get("https://jr.synergy.ru")

# Ввод логина
username_field = driver.find_element(By.XPATH, '//*[@id="username"]')
username_field.send_keys(login_exl)

# Ввод пароля
password_field = driver.find_element(By.XPATH, '//*[@id="password"]')
password_field.send_keys(password_exl)

# Кнопка Входа
login_button = driver.find_element(By.XPATH, '//*[@id="kc-login"]')
login_button.click()

# Перебор всех строк на листе jira
for row in range(2, sheet_jr.max_row + 1):  # Начнем со строки 2
    jr_exl = sheet_jr.cell(row=row, column=1).value  # Чтение ключа задачи
    driver.get(f"https://jr.synergy.ru/browse/{jr_exl}")

    # Ожидание появления комментариев
    try:
        comments_section = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="activitymodule"]/a'))
        )
        comments_section.click()  # Переход к разделу комментариев

        # Здесь Вы можете добавить дополнительный код для работы с комментариями
        time.sleep(7)  # Даем время странице загрузиться

    except Exception as e:
        print(f"Не удалось получить XPATH комментарии для задачи {jr_exl}: {e}")

# Закрыть браузер
input("Проверка ссылок JIRA завершена. Нажмите Enter для закрытия окна")
driver.quit()

# softy_plug